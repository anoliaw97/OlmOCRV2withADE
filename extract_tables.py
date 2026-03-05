"""
extract_tables.py
-----------------
1. Parse all HTML <table> blocks from the two RESULTS text files.
2. Write structured Excel workbooks (one per document).
3. Return python dicts of the key named tables for PPT use.
"""

import re, io
from bs4 import BeautifulSoup

# ── Helpers ───────────────────────────────────────────────────────────────────

def clean(s):
    """Strip LaTeX, extra whitespace, newlines."""
    s = re.sub(r'\\\(.*?\\\)', '', s, flags=re.DOTALL)   # remove \( ... \)
    s = re.sub(r'\\\[.*?\\\]', '', s, flags=re.DOTALL)   # remove \[ ... \]
    return ' '.join(s.split())


def flatten_table(soup_table):
    """
    Convert an HTML <table> (may have rowspan/colspan) into a 2-D list of strings.
    Uses a grid-fill approach to respect rowspans.
    """
    rows_el = soup_table.find_all('tr')
    # First pass: find max cols
    grid = []
    spans = {}   # (row, col) -> remaining span rows

    for ri, tr in enumerate(rows_el):
        grid.append([])
        ci = 0
        for cell in tr.find_all(['th', 'td']):
            # Skip cells occupied by row-spans from above
            while (ri, ci) in spans:
                grid[ri].append(spans.pop((ri, ci)))
                ci += 1
            text = clean(cell.get_text())
            rs = int(cell.get('rowspan', 1))
            cs = int(cell.get('colspan', 1))
            # Fill colspan
            for c in range(cs):
                grid[ri].append(text)
                # Register rowspan for subsequent rows
                for r in range(1, rs):
                    spans[(ri + r, ci + c)] = text
            ci += cs
        # Fill remaining row-span placeholders
        remaining_ci = ci
        while (ri, remaining_ci) in spans:
            grid[ri].append(spans.pop((ri, remaining_ci)))
            remaining_ci += 1

    # Pad all rows to same width
    width = max((len(r) for r in grid), default=0)
    for row in grid:
        while len(row) < width:
            row.append('')
    return grid


def extract_tables_from_file(filepath):
    """
    Parse a RESULTS text file; return list of dicts:
      { 'page': int, 'label': str, 'grid': [[...]] }
    page_label is the text immediately after the </table> tag (if any).
    """
    with open(filepath, encoding='utf-8') as f:
        raw = f.read()

    results = []
    current_page = 0

    # Split by page headers
    page_pattern = re.compile(
        r'={60}\nPAGE (\d+).*?\n={60}(.*?)(?=={60}\nPAGE |\Z)',
        re.DOTALL
    )
    for m in page_pattern.finditer(raw):
        current_page = int(m.group(1))
        block = m.group(2)

        # Find all <table>...</table> blocks in this page
        tbl_pattern = re.compile(r'<table>(.*?)</table>', re.DOTALL)
        pos = 0
        for tm in tbl_pattern.finditer(block):
            soup = BeautifulSoup('<table>' + tm.group(1) + '</table>', 'html.parser')
            tbl = soup.find('table')
            grid = flatten_table(tbl)
            if not grid:
                continue

            # Look for a label in the text after the </table>
            after = block[tm.end():tm.end()+200]
            label_m = re.search(r'^\s*(TABLE\s+[\w.]+[^\n]*)', after, re.MULTILINE | re.IGNORECASE)
            label = clean(label_m.group(1)) if label_m else f'Page {current_page} Table {len(results)+1}'

            results.append({'page': current_page, 'label': label, 'grid': grid})

    return results


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_excel(tables, outpath, sheet_names=None):
    """Write list of table dicts to an Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY_HEX  = '00002060'
    BLUE_HEX  = '004472C4'
    LGRAY_HEX = 'FFF2F2F2'

    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFFFF', size=11)
    hdr_fill  = PatternFill(fill_type='solid', fgColor=NAVY_HEX)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    body_font  = Font(name='Calibri', size=10)
    alt_fill   = PatternFill(fill_type='solid', fgColor=LGRAY_HEX)
    body_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    thin = Side(style='thin', color='FFD6DCE4')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_names = {}
    for idx, t in enumerate(tables):
        # Build safe sheet name (max 31 chars)
        if sheet_names and idx < len(sheet_names):
            sname = sheet_names[idx][:31]
        else:
            raw = t['label'][:28]
            sname = re.sub(r'[\\/*?\[\]:]', '_', raw)
        # Deduplicate
        if sname in used_names:
            used_names[sname] += 1
            sname = sname[:28] + f'_{used_names[sname]}'
        else:
            used_names[sname] = 1

        ws = wb.create_sheet(title=sname)
        # Title row (page label)
        ws.append([f"Source: {t['label']}  (Page {t['page']})"])
        ws.cell(1, 1).font = Font(name='Calibri', bold=True, size=12, color=NAVY_HEX[2:])
        ws.append([])  # blank row

        grid = t['grid']
        for ri, row in enumerate(grid):
            ws.append(row)
            for ci, _ in enumerate(row, 1):
                cell = ws.cell(ri + 3, ci)
                if ri == 0:
                    cell.font  = hdr_font
                    cell.fill  = hdr_fill
                    cell.alignment = hdr_align
                else:
                    cell.font  = body_font
                    cell.fill  = alt_fill if ri % 2 == 0 else PatternFill()
                    cell.alignment = body_align
                cell.border = border

        # Auto-width (rough)
        for ci in range(1, (max(len(r) for r in grid) if grid else 1) + 1):
            max_len = 0
            for row in ws.iter_rows(min_row=3, min_col=ci, max_col=ci):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 10), 30)

        ws.freeze_panes = 'A4'   # freeze header

    wb.save(outpath)
    print(f'  Saved → {outpath}  ({len(wb.sheetnames)} sheets)')
    return wb


# ── Named-table lookup helpers ────────────────────────────────────────────────

def find_table(tables, page=None, label_contains=None):
    for t in tables:
        if page and t['page'] != page:
            continue
        if label_contains and label_contains.lower() not in t['label'].lower():
            continue
        return t
    return None


def grid_to_rows(grid, skip_header_rows=1):
    """Return (headers_list, data_rows_list)."""
    if not grid:
        return [], []
    headers = grid[0] if skip_header_rows >= 1 else []
    data    = grid[skip_header_rows:]
    return headers, data


if __name__ == '__main__':
    DUYONG_TXT = 'RESULTS/Duyon Deep 1 Full.txt'
    PEGAGA_TXT = 'RESULTS/pegaga results.txt'

    print('Parsing Duyong Deep 1...')
    duyong_tables = extract_tables_from_file(DUYONG_TXT)
    print(f'  Found {len(duyong_tables)} tables')
    for t in duyong_tables:
        print(f"    Page {t['page']:2d}: {t['label'][:70]}  [{len(t['grid'])}r x {len(t['grid'][0]) if t['grid'] else 0}c]")

    print()
    print('Parsing Pegaga-2...')
    pegaga_tables = extract_tables_from_file(PEGAGA_TXT)
    print(f'  Found {len(pegaga_tables)} tables')
    for t in pegaga_tables:
        print(f"    Page {t['page']:2d}: {t['label'][:70]}  [{len(t['grid'])}r x {len(t['grid'][0]) if t['grid'] else 0}c]")

    print()
    write_excel(duyong_tables, 'RESULTS/Duyong_Deep1_Extracted_Tables.xlsx')
    write_excel(pegaga_tables, 'RESULTS/Pegaga2_Extracted_Tables.xlsx')
    print('Done.')
