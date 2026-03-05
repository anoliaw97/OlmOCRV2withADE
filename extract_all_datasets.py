"""
extract_all_datasets.py
Extracts ALL structured data from the three OCR text files and writes one
ML-ready Excel workbook per PDF:
  RESULTS/Angsi1_ML_Dataset.xlsx
  RESULTS/DuyongDeep1_ML_Dataset.xlsx
  RESULTS/Pegaga2_ML_Dataset.xlsx
"""

import re, os, math
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── paths ────────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
RES  = os.path.join(BASE, "RESULTS")

# ── Excel style helpers ───────────────────────────────────────────────────────
NAVY  = "002060"; WHITE = "FFFFFF"; LGRAY = "F2F2F2"
LBLUE = "D9E1F2"; YLLOW = "FFF2CC"; GREEN = "E2EFDA"
DKRED = "C00000"

def _fill(h): return PatternFill("solid", fgColor=h)
def _bdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def _font(bold=False, color="000000", size=9, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
def _ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def write_hdr(ws, headers, row=1, bg=NAVY, fg="FFFFFF", height=22):
    ws.row_dimensions[row].height = height
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill      = _fill(bg)
        cell.font      = _font(True, fg, 9)
        cell.alignment = _ctr()
        cell.border    = _bdr()

def write_row(ws, vals, row, bg=WHITE, bold=False):
    ws.row_dimensions[row].height = 15
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill      = _fill(bg)
        cell.font      = _font(bold)
        cell.border    = _bdr()
        cell.alignment = _lft()

def autofit(ws, mn=8, mx=32):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(mn, min(mx, w+3))

def section_row(ws, row, ncols, label, bg=LBLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=label)
    c.fill = _fill(bg); c.font = _font(True, NAVY, 9)
    c.border = _bdr(); c.alignment = _lft()

def safe_num(s):
    """Try to convert to float; return None if not possible."""
    if s is None: return None
    s = str(s).strip().replace(",","").replace("~","").replace("*","")
    try: return float(s)
    except: return s if s else None

# ── HTML table parser ─────────────────────────────────────────────────────────
def clean(s):
    s = re.sub(r'\\\(.*?\\\)', '', s, flags=re.DOTALL)
    s = re.sub(r'\\\[.*?\\\]', '', s, flags=re.DOTALL)
    s = re.sub(r'\*+$', '', s)
    return ' '.join(s.split()).strip()

def flatten(soup_table):
    spans = {}; grid = []; ri = 0
    for tr in soup_table.find_all("tr"):
        cols = tr.find_all(["th","td"]); ci = 0; cells = []; idx = 0
        while idx < len(cols) or ci < 60:
            while (ri, ci) in spans:
                cells.append(spans[(ri, ci)]); ci += 1
            if idx >= len(cols): break
            td  = cols[idx]; idx += 1
            txt = clean(td.get_text())
            rs  = int(td.get("rowspan", 1)); cs = int(td.get("colspan", 1))
            for dr in range(rs):
                for dc in range(cs):
                    if dr == 0 and dc == 0: continue
                    spans[(ri+dr, ci+dc)] = txt
            cells.append(txt); ci += cs
        grid.append(cells); ri += 1
    return grid

def extract_tables_from_text(body):
    """Extract all HTML tables from a body of text, with context."""
    tbls = []
    for m in re.finditer(r'<table[^>]*>(.*?)</table>', body, re.DOTALL|re.IGNORECASE):
        ctx_start = max(0, m.start()-600)
        ctx = body[ctx_start:m.start()]
        soup = BeautifulSoup(m.group(0), "html.parser")
        t = soup.find("table")
        if t:
            tbls.append({"grid": flatten(t), "context": ctx, "raw": m.group(0)})
    return tbls


def parse_pages(filepath):
    """
    Split file into pages. Handles two formats:
    1. OlmOCR format: ={60} PAGE X | Xs | Y tokens ={60}
    2. Simple format: sections separated by --- with front-matter blocks
    Returns list of dicts: {page, dur_s, tokens, text, tables}
    """
    raw = open(filepath, encoding="utf-8").read()

    # Try format 1: PAGE X | dur | tokens markers
    parts = re.split(r'={60,}\s*\nPAGE\s+(\d+)\s+\|\s+([\d.~,]+)s\s+\|\s+(\d+)\s+tokens\s*\n={60,}',
                     raw)
    if len(parts) > 1:
        pages = []
        i = 1
        while i < len(parts):
            pno  = int(parts[i])
            dur  = parts[i+1]
            tok  = int(parts[i+2])
            body = parts[i+3] if i+3 < len(parts) else ""
            pages.append({"page": pno, "dur_s": dur, "tokens": tok,
                          "text": body, "tables": extract_tables_from_text(body)})
            i += 4
        return pages

    # Format 2: sections split by front-matter blocks (--- ... ---)
    # Treat the entire file as one "page" but split by --- blocks
    blocks = re.split(r'^---\s*$', raw, flags=re.MULTILINE)
    # Every block is either front-matter (has primary_language:) or content
    content_blocks = []
    for b in blocks:
        b = b.strip()
        if not b or "primary_language:" in b: continue
        if b.startswith("Attached is one page") or b.startswith("Page ") or \
           b.startswith("TOTAL"): continue
        content_blocks.append(b)

    # Treat as one big page
    full_text = "\n".join(content_blocks)
    return [{"page": 1, "dur_s": "0", "tokens": 0,
             "text": full_text,
             "tables": extract_tables_from_text(full_text)}]

# ─────────────────────────────────────────────────────────────────────────────
# ANGSI-1
# ─────────────────────────────────────────────────────────────────────────────
def extract_angsi(filepath):
    pages = parse_pages(filepath)
    all_tables = [(p["page"], t) for p in pages for t in p["tables"]]

    # TABLE I: Summary of Porosity and Permeability Data
    # Headers: Depth Ft. | Plug No. | Porosity % | Ka_bulk md | Ka_grain md | Bulk dens | Grain dens
    # (row 0 = multi-header, row 1 = sub-header Bulk/Grain, rows 2+ = data)
    t1_grid = all_tables[0][1]["grid"]
    perm_data = []
    for row in t1_grid[2:]:   # skip the two header rows
        if len(row) < 4: continue
        # TABLE I cols: Depth | Plug No. | Porosity% | Ka_bulk md | Bulk Dens | Grain Dens
        # Header sub-row has "Bulk/Grain" for both Perm and Density, but only Ka_bulk
        # and both density columns appear in data (6 values total).
        perm_data.append({
            "Well":             "Angsi-1",
            "Sample_ID":        str(row[1]).strip() if len(row) > 1 else None,
            "Depth_ft":         safe_num(row[0]),
            "Phi_pct":          safe_num(row[2]),
            "Ka_bulk_mD":       safe_num(row[3]),
            "Bulk_Density_gcc": safe_num(row[4]) if len(row) > 4 else None,
            "Grain_Density_gcc":safe_num(row[5]) if len(row) > 5 else None,
        })

    # TABLE II: Gas-Oil Capillary Pressure
    # Structure: Pc_AN30 | Sw_AN30 | Pc_AN36 | Sw_AN36 | (Pc_AN37/AN39 columns empty)
    t2_grid = all_tables[1][1]["grid"]
    pc_data = []
    CORES = [("AN-30", 8015), ("AN-36", 8037)]
    run_labels = {}
    run_idx = 0
    for row in t2_grid[2:]:   # skip 2 header rows
        if len(row) < 4: continue
        # Each row: Pc1, Sw1, Pc2, Sw2
        for ci, (core, depth) in enumerate(CORES):
            pc_val  = safe_num(row[ci*2])    if ci*2 < len(row) else None
            sw_val  = safe_num(row[ci*2+1])  if ci*2+1 < len(row) else None
            if pc_val is None: continue
            raw_pc = str(row[ci*2]).strip()
            # Run 1 = rows 0-4 (5 pts), Run 2 = rows 5-9 (5 pts per core)
            run_no = 1 if run_idx < 5 else 2
            pc_data.append({
                "Well":       "Angsi-1",
                "Sample_ID":  core,
                "Depth_ft":   depth,
                "Pc_psi":     abs(pc_val) if pc_val else None,
                "Sw_pct_PV":  sw_val,
                "Sw_frac":    round(sw_val/100, 4) if sw_val else None,
                "Estimated":  "*" in raw_pc or "1" in raw_pc[:2],
                "Run":        run_no,
                "Fluid_System":"Gas-Oil (Air-Kerosene)",
                "IFT_dynes_cm":28.0,
            })
        run_idx += 1  # advance per row pair (same row_idx for both cores)

    # Fix run_idx – it's per-data-row, not per-core; so use enumerate
    pc_data = []
    for ri, row in enumerate(t2_grid[2:]):
        if len(row) < 4: continue
        for ci, (core, depth) in enumerate(CORES):
            pc_raw  = row[ci*2]   if ci*2 < len(row) else ""
            sw_raw  = row[ci*2+1] if ci*2+1 < len(row) else ""
            pc_val  = safe_num(pc_raw);  sw_val = safe_num(sw_raw)
            if pc_val is None: continue
            pc_data.append({
                "Well":        "Angsi-1",
                "Sample_ID":   core,
                "Depth_ft":    depth,
                "Run":         1 if ri < 5 else 2,
                "Pc_psi":      abs(float(pc_val)) if pc_val else None,
                "Sw_pct_PV":   float(sw_val) if sw_val else None,
                "Sw_frac":     round(float(sw_val)/100, 4) if sw_val else None,
                "log10_Pc":    round(math.log10(abs(float(pc_val))),4) if pc_val else None,
                "Estimated":   str(pc_raw).strip().endswith("*"),
                "Fluid":       "Gas-Oil",
                "Method":      "Centrifuge",
                "IFT_dynes_cm":28.0,
            })

    # TABLE III: Electrical Properties  (pivoted – rows are properties, cols are samples)
    # grid[0] = [" ", "Water Sat % PV", "", "Resistivity Ratio R/Ro"]
    # Then blocks of rows: Core No, Depth ft, Porosity %, Permeability md, FF, Sat_Exp
    t3_grid = all_tables[2][1]["grid"]
    # Parse pivot: each ~6-row block = one sample
    elec_data = []
    cur = {}
    for row in t3_grid:
        if not row: continue
        key = str(row[0]).strip()
        if key == "Core No.":
            if cur: elec_data.append(cur)
            cur = {"Well":"Angsi-1","Sample_ID": str(row[1]).strip() if len(row)>1 else None}
            # Additional columns = water sat values (for resistivity ratio data)
            cur["_extra_sw"]  = safe_num(row[2]) if len(row)>2 else None
            cur["_extra_rr"]  = safe_num(row[3]) if len(row)>3 else None
        elif key == "Depth, ft.":
            cur["Depth_ft"] = safe_num(row[1]) if len(row)>1 else None
        elif key == "Porosity, %":
            cur["Phi_pct"]   = safe_num(row[1]) if len(row)>1 else None
            cur["Sw2_pct"]   = safe_num(row[2]) if len(row)>2 else None
            cur["RR2"]       = safe_num(row[3]) if len(row)>3 else None
        elif key == "Permeability, md":
            cur["Ka_mD"]    = safe_num(row[1]) if len(row)>1 else None
        elif key == "Formation Factor":
            cur["Formation_Factor"] = safe_num(row[1]) if len(row)>1 else None
        elif key == "Saturation Exponent":
            cur["Sat_Exponent"] = safe_num(row[1]) if len(row)>1 else None
    if cur: elec_data.append(cur)
    # Clean up internal temp keys
    for d in elec_data:
        d.pop("_extra_sw", None); d.pop("_extra_rr", None)
        d.pop("Sw2_pct", None);   d.pop("RR2", None)

    # Merge perm + electrical
    elec_map = {d["Sample_ID"]: d for d in elec_data}
    full_props = []
    for p in perm_data:
        row = dict(p)
        sid = p["Sample_ID"]
        e   = elec_map.get(sid, {})
        row["Formation_Factor"]  = e.get("Formation_Factor")
        row["Sat_Exponent"]      = e.get("Sat_Exponent")
        row["RQI"]               = round(0.0314*math.sqrt(p["Ka_bulk_mD"]/
                                   (p["Phi_pct"]/100)),4) if p["Ka_bulk_mD"] and p["Phi_pct"] else None
        row["log10_Ka"]          = round(math.log10(p["Ka_bulk_mD"]),4) if p["Ka_bulk_mD"] else None
        full_props.append(row)

    return {"sample_properties": full_props, "capillary_pressure": pc_data,
            "electrical": elec_data}


def write_angsi(data, outpath):
    wb = openpyxl.Workbook()

    # ── Single sheet: All_Features_ML ────────────────────────────────────────
    ws = wb.active; ws.title = "ML_Dataset"
    all_hdrs = ["Well","Sample_ID","Depth_ft","Phi_pct","Ka_bulk_mD",
                "Bulk_Density_gcc","Grain_Density_gcc","Formation_Factor","Sat_Exponent",
                "RQI","log10_Ka",
                "Pc_entry_psi_Run1","Sw_at_Pc_entry_Run1",
                "Pc_max_psi_Run1","Sw_at_Pc_max_Run1",
                "Pc_entry_psi_Run2","Sw_at_Pc_entry_Run2",
                "Pc_max_psi_Run2","Sw_at_Pc_max_Run2"]
    write_hdr(ws, all_hdrs)
    pc_map = {}
    for d in data["capillary_pressure"]:
        sid = d["Sample_ID"]; run = d["Run"]
        pc_map.setdefault(sid, {}).setdefault(run, []).append(d)
    for ri, d in enumerate(data["sample_properties"], 2):
        sid = d["Sample_ID"]
        row_vals = [d.get(h) for h in all_hdrs[:12]]
        for run in [1, 2]:
            pts = sorted(pc_map.get(sid, {}).get(run, []), key=lambda x: x["Pc_psi"] or 0)
            if pts:
                entry = pts[0]; maxx = pts[-1]
                row_vals += [entry["Pc_psi"], entry["Sw_pct_PV"],
                             maxx["Pc_psi"],  maxx["Sw_pct_PV"]]
            else:
                row_vals += [None, None, None, None]
        bg = LGRAY if ri%2==0 else WHITE
        write_row(ws, row_vals, ri, bg)
    ws.freeze_panes = "A2"; autofit(ws)

    wb.save(outpath)
    print(f"  Saved: {outpath}  ({sum(1 for _ in data['sample_properties'])} samples)")


# ─────────────────────────────────────────────────────────────────────────────
# DUYONG DEEP-1
# ─────────────────────────────────────────────────────────────────────────────
def classify_duyong_table(grid, context):
    """Return a label for the table based on header structure and context."""
    if not grid: return "unknown"
    hdr = " ".join(str(c) for c in grid[0]).lower()
    ctx = context.lower()

    if "sample id" in hdr and "frf" in hdr.replace(" ",""):
        return "FRF_summary"
    if "sample id" in hdr and "resistivity" in hdr:
        return "resistivity_index"
    if "sample id" in hdr and "depth" in hdr and "cec" in hdr:
        return "CEC"
    if "sample id" in hdr and "depth" in hdr and "porosity" in hdr and "grain" in hdr:
        return "cap_press_summary"
    if "injection pressure" in hdr or "mercury saturation" in hdr:
        return "MICP"
    if "pore size" in hdr or "pore throat" in hdr or ("equivalent diameter" in hdr):
        return "pore_size"
    if ("sw" in hdr or "water sat" in hdr) and ("resistivity" in hdr or "frf" in hdr):
        return "FRF_detail"
    if "sample id" in hdr and "depth" in hdr and "frf" in ctx:
        return "FRF_summary"
    if "frf" in ctx or "formation resistivity" in ctx:
        return "FRF_detail"
    # Pore size distribution / conversion tables
    if len(grid) > 1 and len(grid[0]) >= 4 and ("psi" in hdr or "radius" in hdr):
        return "MICP"
    return "other"


def extract_duyong(filepath):
    pages = parse_pages(filepath)

    frf_summary = []    # Table 1
    frf_detail  = []    # Table 2 (multiple samples)
    cap_summary = []    # Table 3
    cec_data    = []    # Table 4
    micp_data   = []    # Table 5A-5D (per sample)
    pore_size   = []    # pore size distribution tables

    # Track current sample from page text
    current_sample = None
    current_depth  = None

    for page in pages:
        # Detect sample context from page text
        m = re.search(r'Sample ID\s*[:\s]+(\d+)\s*\nDepth\s+([0-9.]+)\s*meter', page["text"])
        if m:
            current_sample = int(m.group(1))
            current_depth  = float(m.group(2))
        # Also check for TABLE 5 label
        m2 = re.search(r'TABLE 5[A-D].*?\n.*?Sample ID\s*(\d+)\s*\nDepth\s+([0-9.]+)', page["text"], re.DOTALL)
        if m2:
            current_sample = int(m2.group(1))
            current_depth  = float(m2.group(2))

        for tbl in page["tables"]:
            grid = tbl["grid"]; ctx = tbl["context"]
            if not grid or not grid[0]: continue

            label = classify_duyong_table(grid, ctx)
            hdr   = [str(c).strip() for c in grid[0]]
            hdr2  = [str(c).strip() for c in grid[1]] if len(grid) > 1 else []

            if label == "FRF_summary":
                # Rows: Sample ID, Depth, Phi, Grain dens, ...
                for row in grid[1:]:
                    if len(row) < 3: continue
                    frf_summary.append({
                        "Well":        "Duyong Deep-1",
                        "Sample_ID":   f"D-{safe_num(row[0]):03.0f}" if safe_num(row[0]) else row[0],
                        "Depth_m":     safe_num(row[1]),
                        "Phi_pct":     safe_num(row[2]),
                        "Grain_Density_gcc": safe_num(row[3]) if len(row)>3 else None,
                        "col4":        safe_num(row[4]) if len(row)>4 else None,
                        "col5":        safe_num(row[5]) if len(row)>5 else None,
                    })

            elif label == "resistivity_index":
                for row in grid[2:]:  # skip 2 header rows
                    if len(row) < 3: continue
                    sid_raw = safe_num(row[0])
                    sid_str = f"D-{int(sid_raw):03d}" if isinstance(sid_raw, float) else str(row[0])
                    frf_detail.append({
                        "Well":       "Duyong Deep-1",
                        "Sample_ID":  sid_str,
                        "Depth_m":    safe_num(row[1]),
                        "Phi_pct":    safe_num(row[2]),
                        "Ka_mD":      safe_num(row[3]) if len(row)>3 else None,
                        "col4":       safe_num(row[4]) if len(row)>4 else None,
                        "col5":       safe_num(row[5]) if len(row)>5 else None,
                        "col6":       safe_num(row[6]) if len(row)>6 else None,
                        "col7":       safe_num(row[7]) if len(row)>7 else None,
                        "col8":       safe_num(row[8]) if len(row)>8 else None,
                    })

            elif label == "CEC":
                for row in grid[1:]:
                    if len(row) < 3: continue
                    sid_raw = safe_num(row[0])
                    sid_str = f"D-{int(sid_raw):03d}" if isinstance(sid_raw, float) else str(row[0])
                    cec_data.append({
                        "Well":            "Duyong Deep-1",
                        "Sample_ID":       sid_str,
                        "Depth_m":         safe_num(row[1]),
                        "Phi_pct":         safe_num(row[2]),
                        "Grain_Density_gcc":safe_num(row[3]) if len(row)>3 else None,
                        "CEC_meq100g":     safe_num(row[4]) if len(row)>4 else None,
                        "Qv_meq_ml":       safe_num(row[5]) if len(row)>5 else None,
                    })

            elif label == "cap_press_summary":
                for row in grid[1:]:
                    if len(row) < 3: continue
                    sid_raw = safe_num(row[0])
                    sid_str = f"D-{int(sid_raw):03d}" if isinstance(sid_raw, float) else str(row[0])
                    cap_summary.append({
                        "Well":      "Duyong Deep-1",
                        "Sample_ID": sid_str,
                        "Depth_m":   safe_num(row[1]),
                        "Phi_pct":   safe_num(row[2]),
                        "Grain_Density_gcc": safe_num(row[3]) if len(row)>3 else None,
                        "CEC_meq100g":       safe_num(row[4]) if len(row)>4 else None,
                        "Qv_meq_ml":         safe_num(row[5]) if len(row)>5 else None,
                    })

            elif label == "MICP":
                if current_sample is None: continue
                sid_str = f"D-{current_sample:03d}"
                # Header rows describe columns – use indices based on column count
                # Cols: Pc_psia | Hg_Sat_fracVp | Hg_Sat_fracVb | PoreThrad_AB | PT_OB | ... | Sw_fracVp | J_Function
                ncols = len(grid[0]) + len(grid[1] if len(grid)>1 else [])
                start_row = 2 if (len(grid)>1 and not safe_num(grid[1][0])) else 1
                for row in grid[start_row:]:
                    if len(row) < 4: continue
                    pc = safe_num(row[0])
                    if not isinstance(pc, (int, float)) or pc is None: continue
                    hg_sat_vp = safe_num(row[1]) if len(row)>1 else None
                    hg_sat_vb = safe_num(row[2]) if len(row)>2 else None
                    pt_ab     = safe_num(row[3]) if len(row)>3 else None
                    pt_ob_lab = safe_num(row[4]) if len(row)>4 else None
                    pt_ob_res = safe_num(row[5]) if len(row)>5 else None
                    eqpc_ab   = safe_num(row[6]) if len(row)>6 else None  # or next cols
                    eqpc_ob_lab = safe_num(row[7]) if len(row)>7 else None
                    eqpc_ob_res = safe_num(row[8]) if len(row)>8 else None
                    sw_frac   = safe_num(row[-2]) if len(row)>=2 else None
                    j_func    = safe_num(row[-1]) if len(row)>=1 else None
                    micp_data.append({
                        "Well":              "Duyong Deep-1",
                        "Sample_ID":         sid_str,
                        "Depth_m":           current_depth,
                        "Pc_inj_psia":       pc,
                        "Hg_Sat_fracVp":     hg_sat_vp,
                        "Hg_Sat_fracVb":     hg_sat_vb,
                        "PoreThroat_AB_micron":   pt_ab,
                        "PoreThroat_OB_lab_micron":pt_ob_lab,
                        "PoreThroat_OB_res_micron":pt_ob_res,
                        "Sw_wetting_fracVp": sw_frac,
                        "J_Function":        j_func,
                        "log10_Pc":          round(math.log10(pc),4) if pc and pc>0 else None,
                    })

    # Remove duplicate FRF rows (same sample may appear in multiple table types)
    seen = set()
    frf_clean = []
    for d in frf_detail:
        key = (d["Sample_ID"], d.get("Depth_m"))
        if key not in seen:
            seen.add(key); frf_clean.append(d)

    # Identify unique TABLE 1 rows properly
    # TABLE 1 structure: Sample ID | Depth | Phi | col3 | col4 | col5 | ...
    # Rebuild with actual FRF header names from the text
    # TABLE 1 headers: Sample ID | Depth meters | Ka Air mD | Phi % | FRF_NOB | RI_NOB | ...
    # We need to re-read TABLE 1 specifically
    frf_t1 = _parse_duyong_t1(filepath)
    frf_t2 = _parse_duyong_t2(filepath)

    frf_t1   = _parse_duyong_t1(filepath)   # FRF + RI long format
    airbrine = _parse_duyong_t2(filepath)   # Air-Brine Centrifuge Pc

    return {"frf_ri_long": frf_t1, "airbrine_pc": airbrine, "cec": cec_data,
            "micp": micp_data, "cap_summary": cap_summary}


def _get_table_after_label(text, label_pattern):
    """Extract the first <table> after a line matching label_pattern."""
    m = re.search(label_pattern + r'.*?(<table.*?</table>)', text,
                  re.DOTALL | re.IGNORECASE)
    if not m: return []
    soup = BeautifulSoup(m.group(1), "html.parser")
    t = soup.find("table")
    return flatten(t) if t else []


def _parse_duyong_t1(filepath):
    """
    TABLE 1: FRF + Resistivity Index at NOB (long-format per sample × Sw step).
    Cols: Sample ID(rowspan), Depth(rowspan), Ka_NOB(rowspan), Phi_NOB(rowspan),
          FRF(rowspan), m_Archie(rowspan), Sw_frac, RI, n_sat_exp
    """
    text = open(filepath).read()
    grid = _get_table_after_label(text, r'TABLE 1:')
    if not grid: return []
    rows = []
    # Rows 0-1 are header rows; rows 2+ are data
    for row in grid[2:]:
        if len(row) < 7: continue
        sid_raw = safe_num(row[0])
        if not isinstance(sid_raw, (int, float)): continue
        sid = f"D-{int(sid_raw):03d}"
        ri_val = safe_num(row[7]) if len(row) > 7 else None
        n_val  = safe_num(row[8]) if len(row) > 8 else None
        rows.append({
            "Well":          "Duyong Deep-1",
            "Sample_ID":     sid,
            "Depth_m":       safe_num(row[1]),
            "Ka_NOB_mD":     safe_num(row[2]),
            "Phi_NOB_frac":  safe_num(row[3]),
            "FRF":           safe_num(row[4]),
            "m_Archie":      safe_num(row[5]),
            "Sw_frac":       safe_num(row[6]),
            "RI":            ri_val if ri_val != "-" else None,
            "n_sat_exp":     n_val  if n_val  != "-" else None,
        })
    return rows


def _parse_duyong_t2(filepath):
    """
    TABLE 2: Air-Brine Centrifuge Capillary Pressure (Brine Saturation at various Pc).
    Cols: Sample ID, Depth, NOB_psi, Ka_air_mD, Phi_pct,
          Sw@Pc=0, Sw@Pc=1, Sw@Pc=2, Sw@Pc=4, Sw@Pc=8,
          Sw@Pc=15, Sw@Pc=25, Sw@Pc=50, Sw@Pc=100, Sw@Pc=200 psi
    Returns long-format: one row per (sample, Pc_level).
    """
    text = open(filepath).read()
    grid = _get_table_after_label(text, r'TABLE 2:')
    if not grid: return []
    # The 10 Pc pressure levels (from sub-header row)
    PC_LEVELS = [0, 1, 2, 4, 8, 15, 25, 50, 100, 200]
    rows = []
    for row in grid[2:]:   # skip 2 header rows
        if len(row) < 6: continue
        sid_raw = safe_num(row[0])
        if not isinstance(sid_raw, (int, float)): continue
        sid = f"D-{int(sid_raw):03d}"
        depth = safe_num(row[1]); nob = safe_num(row[2])
        ka    = safe_num(row[3]); phi = safe_num(row[4])
        for pc_i, pc_psi in enumerate(PC_LEVELS):
            col = 5 + pc_i
            sw  = safe_num(row[col]) if col < len(row) else None
            if sw is None: continue
            rows.append({
                "Well":      "Duyong Deep-1",
                "Sample_ID": sid,
                "Depth_m":   depth,
                "NOB_psi":   nob,
                "Ka_air_mD": ka,
                "Phi_pct":   phi,
                "Pc_psi":    pc_psi,
                "Sw_brine_pct_PV": sw,
                "Sw_frac":   round(sw / 100, 4) if sw else None,
            })
    return rows


def write_duyong(data, outpath):
    wb = openpyxl.Workbook()
    sample_colors = {"D-002":WHITE,"D-003":LGRAY,"D-004":LBLUE,"D-006":YLLOW}

    # Build props lookup (needed for single sheet)
    cec_map = {d["Sample_ID"]:d for d in data["cec"]}
    frf_map = {}
    for d in data["frf_ri_long"]:
        sid = d["Sample_ID"]
        if sid not in frf_map:
            frf_map[sid] = d
    props = []
    for sid in ["D-002","D-003","D-004","D-006"]:
        cec = cec_map.get(sid, {}); frf = frf_map.get(sid, {})
        ka  = frf.get("Ka_NOB_mD"); phi_frac = frf.get("Phi_NOB_frac")
        phi_pct = round(phi_frac*100, 2) if phi_frac else cec.get("Phi_pct")
        rqi = round(0.0314*math.sqrt(ka/phi_frac),4) if ka and phi_frac else None
        props.append({
            "Well":             "Duyong Deep-1",
            "Sample_ID":        sid,
            "Depth_m":          frf.get("Depth_m") or cec.get("Depth_m"),
            "Phi_pct":          phi_pct,
            "Ka_NOB_mD":        ka,
            "Grain_Density_gcc":cec.get("Grain_Density_gcc"),
            "CEC_meq100g":      cec.get("CEC_meq100g"),
            "Qv_meq_ml":        cec.get("Qv_meq_ml"),
            "FRF":              frf.get("FRF"),
            "m_Archie":         frf.get("m_Archie"),
            "RQI":              rqi,
            "log10_Ka":         round(math.log10(ka),4) if ka else None,
        })

    # ── Single sheet: ML_Dataset (all features per sample) ───────────────────
    ws = wb.active; ws.title = "ML_Dataset"
    from collections import defaultdict
    by_sample = defaultdict(list)
    for d in data["micp"]: by_sample[d["Sample_ID"]].append(d)
    micp_stats = {}
    for sid, rows in by_sample.items():
        pts = [r for r in rows if isinstance(r.get("Hg_Sat_fracVp"),(int,float)) and r["Hg_Sat_fracVp"]>0]
        if pts:
            entry = min(pts, key=lambda x: x["Pc_inj_psia"])
            last  = max(pts, key=lambda x: x["Pc_inj_psia"])
            micp_stats[sid] = {
                "Pc_entry_psia": entry["Pc_inj_psia"],
                "Sw_at_entry":   entry["Sw_wetting_fracVp"],
                "Pc_max_psia":   last["Pc_inj_psia"],
                "Sw_irr_fracVp": last["Sw_wetting_fracVp"],
                "n_micp_pts":    len(rows),
            }
    ab_stats = {}
    for d in data["airbrine_pc"]:
        sid = d["Sample_ID"]
        ab_stats.setdefault(sid, []).append(d)
    all_hdrs = ["Well","Sample_ID","Depth_m","Phi_pct","Ka_NOB_mD","Grain_Density_gcc",
                "CEC_meq100g","Qv_meq_ml","FRF","m_Archie","RQI","log10_Ka",
                "Sw_at_AirBrine_Pc200psi",
                "Pc_entry_MICP_psia","Sw_at_entry_MICP","Pc_max_MICP_psia","Sw_irr_MICP","n_micp_pts"]
    write_hdr(ws, all_hdrs)
    for ri, p in enumerate(props, 2):
        sid = p["Sample_ID"]
        ms  = micp_stats.get(sid, {})
        ab  = {d["Pc_psi"]:d["Sw_frac"] for d in ab_stats.get(sid,[])}
        bg  = sample_colors.get(sid, WHITE)
        write_row(ws, [
            p.get("Well"), p.get("Sample_ID"), p.get("Depth_m"), p.get("Phi_pct"),
            p.get("Ka_NOB_mD"), p.get("Grain_Density_gcc"),
            p.get("CEC_meq100g"), p.get("Qv_meq_ml"), p.get("FRF"), p.get("m_Archie"),
            p.get("RQI"), p.get("log10_Ka"),
            ab.get(200),
            ms.get("Pc_entry_psia"), ms.get("Sw_at_entry"),
            ms.get("Pc_max_psia"),   ms.get("Sw_irr_fracVp"), ms.get("n_micp_pts"),
        ], ri, bg)
    ws.freeze_panes="A2"; autofit(ws)

    wb.save(outpath)
    print(f"  Saved: {outpath}  ({len(props)} samples, {len(data['micp'])} MICP pts)")


# ─────────────────────────────────────────────────────────────────────────────
# PEGAGA-2
# ─────────────────────────────────────────────────────────────────────────────
def classify_pegaga_table(grid, context):
    if not grid or not grid[0]: return "unknown"
    hdr = " ".join(str(c) for c in grid[0]).lower()
    ctx = context.lower()
    if "sample id" in hdr and "achieved water" in hdr:
        return "sample_properties"
    # kr curve: must check BEFORE params (same Swn/Sw headers)
    if "swn" in hdr and ("krw" in hdr or "krg" in hdr):
        return "kr_curve"
    # imbibition params: has sgr, frac — check BEFORE drainage (both have swir, frac)
    if "sgr, frac" in hdr and "nw" in hdr and "ng" in hdr:
        return "imbibition_params"
    if "swir, frac" in hdr and "nw" in hdr and "ng" in hdr:
        return "drainage_params"
    if "lab data uss" in hdr or ("lab data" in hdr and "krg" in hdr):
        return "kr_lab_comparison"
    if "nw" in hdr and "ng" in hdr and len(grid) <= 3:
        return "normalized_corey"
    return "other"


def _sample_idx(sid):
    """Map 2-015 → 1, 2-019 → 2, etc."""
    return ["2-015","2-019","2-023","2-029","2-033"].index(sid) + 1


def extract_pegaga(filepath):
    """
    Parse Pegaga-2 text file. Identifies tables by:
      - classify_pegaga_table() for structural matching
      - Explicit 'Table 3.X.Y' labels in context (X=sample 1-5, Y=1-4)
        Y=1 → drainage params, Y=2 → drainage kr
        Y=3 → imbibition params, Y=4 → imbibition kr
      - Fallback: track sample from nearest 'Sample 2-0XX' mention,
        and use sequential order (params → kr → params → kr)
    """
    pages = parse_pages(filepath)
    SAMPLE_IDS = ["2-015","2-019","2-023","2-029","2-033"]

    sample_props    = []
    drainage_params = {}
    imb_params      = {}
    kr_drainage     = {}
    kr_imbibition   = {}

    current_sample = None
    # Per sample, track which table type we've seen last (to infer cycle for unlabeled kr tables)
    # Order within each sample: drn_params → drn_kr → imb_params → imb_kr
    sample_stage = {sid: 0 for sid in SAMPLE_IDS}
    # 0=start, 1=seen drn_params, 2=seen drn_kr, 3=seen imb_params, 4=seen imb_kr

    def infer_cycle(sid, label):
        """Return 'drainage' or 'imbibition' based on label and sample stage."""
        stage = sample_stage.get(sid, 0)
        if label == "drainage_params":
            return "drainage"
        if label == "imbibition_params":
            return "imbibition"
        if label == "kr_curve":
            # Use stage to infer
            return "imbibition" if stage >= 2 else "drainage"
        return "drainage"

    # Per-sample last known kr target: 'drainage' or 'imbibition'
    last_kr_cycle   = {sid: None for sid in SAMPLE_IDS}
    last_explicit_sample = None   # last sample with an explicit T3.X.Y label

    for page in pages:
        txt = page["text"]
        # Page-level sample detection (low priority, may span section boundaries)
        page_sample = None
        for sid in SAMPLE_IDS:
            if f"Sample {sid}" in txt or f"Plug {sid}" in txt:
                page_sample = sid; break
        if page_sample: current_sample = page_sample

        for tbl in page["tables"]:
            grid = tbl["grid"]; ctx = tbl["context"]
            if not grid or not grid[0]: continue

            # Table-level sample detection from context (higher priority)
            ctx_sample = None
            for sid in SAMPLE_IDS:
                if sid in ctx:
                    ctx_sample = sid; break
            if ctx_sample: current_sample = ctx_sample

            # Explicit Table 3.X.Y label — use the LAST (closest) match in context
            all_matches = list(re.finditer(r'Table\s+3\.(\d+)[\s.]+(\d+)', ctx, re.IGNORECASE))
            explicit_sample = None; explicit_tbl_type = None; explicit_y = None
            if all_matches:
                m = all_matches[-1]   # use closest match to table
                x = int(m.group(1)); y = int(m.group(2))
                if 1 <= x <= 5:
                    explicit_sample = SAMPLE_IDS[x-1]
                    explicit_y = y
                    if y == 1:   explicit_tbl_type = "drainage_params"
                    elif y == 2: explicit_tbl_type = "drainage_kr"
                    elif y == 3: explicit_tbl_type = "imbibition_params"
                    elif y == 4: explicit_tbl_type = "imbibition_kr"
            if explicit_sample:
                current_sample = explicit_sample
                last_explicit_sample = explicit_sample

            label = classify_pegaga_table(grid, ctx)

            # Determine cycle with priority:
            #   1. explicit label Y=2 or Y=4 → direct kr cycle assignment
            #   2. structural label + explicit context: Y=1 context + kr_curve → drainage kr
            #      (the kr data immediately follows the drainage params on same page)
            #      Y=3 context + kr_curve → imbibition kr (first imbibition kr rows)
            #   3. No explicit → continuation table, use last_kr_cycle
            if label == "kr_curve":
                if explicit_tbl_type == "drainage_kr":
                    cycle = "drainage"
                elif explicit_tbl_type == "imbibition_kr":
                    cycle = "imbibition"
                elif explicit_tbl_type == "drainage_params" and explicit_y == 1:
                    # kr table immediately after drainage params → this IS drainage kr (T3.X.2)
                    cycle = "drainage"
                elif explicit_tbl_type == "imbibition_params" and explicit_y == 3:
                    # kr table after imbibition params context → this is start of imbibition kr
                    cycle = "imbibition"
                else:
                    # Continuation table: no explicit label, no context match
                    # Use last_explicit_sample to avoid cross-sample contamination
                    cont_sid = last_explicit_sample if last_explicit_sample else current_sample
                    if cont_sid and last_kr_cycle[cont_sid] is not None:
                        cycle = last_kr_cycle[cont_sid]
                        current_sample = cont_sid  # stay with last explicit sample
                    else:
                        cycle = infer_cycle(current_sample, label) if current_sample else "drainage"
                if current_sample: last_kr_cycle[current_sample] = cycle
            elif explicit_tbl_type == "drainage_params" or label == "drainage_params":
                label = "drainage_params"; cycle = "drainage"
            elif explicit_tbl_type == "imbibition_params" or label == "imbibition_params":
                label = "imbibition_params"; cycle = "imbibition"
            else:
                cycle = infer_cycle(current_sample, label) if current_sample else "drainage"

            # ── Sample Properties ─────────────────────────────────────────────
            if label == "sample_properties":
                for row in grid[1:]:
                    if len(row) < 4: continue
                    sid = str(row[0]).strip()
                    if not re.match(r'2-\d{3}', sid): continue
                    sample_props.append({
                        "Well":            "Pegaga-2",
                        "Sample_ID":       sid,
                        "Depth_m":         safe_num(row[1]),
                        "Ka_mD":           safe_num(row[2]),
                        "Phi_pct":         safe_num(row[3]),
                        "Swi_target_pct":  safe_num(row[4]) if len(row)>4 else None,
                        "Swi_achieved_pct":safe_num(row[5]) if len(row)>5 else None,
                    })

            # ── Drainage params ───────────────────────────────────────────────
            elif label == "drainage_params" and current_sample:
                for row in grid[1:]:
                    if len(row) < 4: continue
                    v = [safe_num(row[i]) if i<len(row) else None for i in range(7)]
                    drainage_params[current_sample] = {
                        "Kw_abs_mD": v[0], "Swir_frac": v[1], "Kg_Swir_mD": v[2],
                        "Krw_init":  v[3], "Krg_Swir":  v[4], "Nw": v[5], "Ng": v[6],
                    }
                    sample_stage[current_sample] = max(sample_stage[current_sample], 1)
                    break

            # ── Imbibition params ─────────────────────────────────────────────
            elif label == "imbibition_params" and current_sample:
                for row in grid[1:]:
                    if len(row) < 5: continue
                    v = [safe_num(row[i]) if i<len(row) else None for i in range(8)]
                    imb_params[current_sample] = {
                        "Kg_Swir_mD": v[0], "Swir_frac": v[1], "Kw_Sgr_mD": v[2],
                        "Sgr_frac":   v[3], "Krg_Swir":  v[4], "Krw_Sgr":   v[5],
                        "Nw":         v[6], "Ng":         v[7],
                    }
                    sample_stage[current_sample] = max(sample_stage[current_sample], 3)
                    break

            # ── Kr curve data ─────────────────────────────────────────────────
            elif label == "kr_curve" and current_sample:
                target = kr_drainage if cycle == "drainage" else kr_imbibition
                target.setdefault(current_sample, [])
                added = 0
                for row in grid[1:]:
                    if len(row) < 4: continue
                    swn = safe_num(row[0])
                    if not isinstance(swn, (int,float)): continue
                    target[current_sample].append({
                        "Swn": swn,
                        "Sw":  safe_num(row[1]),
                        "Krw": safe_num(row[2]),
                        "Sw2": safe_num(row[3]) if len(row)>3 else None,
                        "Krg": safe_num(row[4]) if len(row)>4 else None,
                    })
                    added += 1
                if added > 0:
                    if cycle == "drainage":
                        sample_stage[current_sample] = max(sample_stage[current_sample], 2)
                    else:
                        sample_stage[current_sample] = max(sample_stage[current_sample], 4)

    # De-duplicate and sort all kr data
    def dedup_sort(kr_dict):
        for sid in list(kr_dict.keys()):
            seen = set(); clean = []
            for row in kr_dict[sid]:
                k = row["Swn"]
                if k not in seen:
                    seen.add(k); clean.append(row)
            kr_dict[sid] = sorted(clean, key=lambda x: x["Swn"])

    dedup_sort(kr_drainage)
    dedup_sort(kr_imbibition)

    return {"sample_props": sample_props, "drainage_params": drainage_params,
            "imb_params": imb_params, "kr_drainage": kr_drainage,
            "kr_imbibition": kr_imbibition}


def write_pegaga(data, outpath):
    wb = openpyxl.Workbook()
    SAMPLE_IDS = ["2-015","2-019","2-023","2-029","2-033"]
    sample_colors = {
        "2-015": WHITE, "2-019": LGRAY, "2-023": LBLUE,
        "2-029": YLLOW, "2-033": GREEN,
    }
    sp_map = {d["Sample_ID"]:d for d in data["sample_props"]}

    # ── Single sheet: ML_Dataset (all features per sample) ───────────────────
    ws = wb.active; ws.title = "ML_Dataset"
    ml_hdrs = [
        "Well","Sample_ID","Depth_m","Ka_mD","Phi_pct","RQI","log10_Ka",
        "Swi_achieved_pct",
        "Kw_abs_mD","Swir_frac","Kg_Swir_mD","Nw_drn","Ng_drn","n_drn_pts",
        "Sgr_frac","Kw_Sgr_mD","Krw_at_Sgr","Nw_imb","Ng_imb","n_imb_pts",
        "GasRecovery_pct","Krg_max","Swn_at_crossover",
    ]
    write_hdr(ws, ml_hdrs)
    for ri, sid in enumerate(SAMPLE_IDS, 2):
        sp  = sp_map.get(sid, {})
        drn = data["drainage_params"].get(sid, {})
        imb = data["imb_params"].get(sid, {})
        ka  = sp.get("Ka_mD"); phi = sp.get("Phi_pct")
        rqi  = round(0.0314*math.sqrt(ka/(phi/100)),4) if ka and phi else None
        log_ka = round(math.log10(ka),4) if ka else None
        swir = imb.get("Swir_frac") or drn.get("Swir_frac")
        sgr  = imb.get("Sgr_frac")
        gas_rec = round((1-swir-sgr)/(1-swir)*100,1) if swir and sgr else None
        n_drn = len(data["kr_drainage"].get(sid, []))
        n_imb = len(data["kr_imbibition"].get(sid, []))
        crossover = None
        for row in data["kr_imbibition"].get(sid, []):
            if row.get("Krw") and row.get("Krg"):
                try:
                    if abs(float(row["Krw"]) - float(row["Krg"])) < 0.02:
                        crossover = row["Swn"]; break
                except: pass
        bg = sample_colors.get(sid, WHITE)
        write_row(ws, [
            "Pegaga-2", sid, sp.get("Depth_m"), ka, phi, rqi, log_ka,
            sp.get("Swi_achieved_pct"),
            drn.get("Kw_abs_mD"), drn.get("Swir_frac"), drn.get("Kg_Swir_mD"),
            drn.get("Nw"), drn.get("Ng"), n_drn,
            imb.get("Sgr_frac"), imb.get("Kw_Sgr_mD"), imb.get("Krw_Sgr"),
            imb.get("Nw"), imb.get("Ng"), n_imb,
            gas_rec, 1.0, crossover,
        ], ri, bg)
    ws.freeze_panes="A2"; autofit(ws)

    wb.save(outpath)
    total_kr = sum(len(v) for v in data["kr_drainage"].values()) + \
               sum(len(v) for v in data["kr_imbibition"].values())
    print(f"  Saved: {outpath}  ({len(data['sample_props'])} samples, {total_kr} kr data points)")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    os.makedirs(RES, exist_ok=True)

    print("Extracting Angsi-1 …")
    angsi = extract_angsi(os.path.join(RES, "Angsi 1 Core.txt"))
    write_angsi(angsi, os.path.join(RES, "Angsi1_ML_Dataset.xlsx"))

    print("Extracting Duyong Deep-1 …")
    duyong = extract_duyong(os.path.join(RES, "Duyon Deep 1 Full.txt"))
    write_duyong(duyong, os.path.join(RES, "DuyongDeep1_ML_Dataset.xlsx"))

    print("Extracting Pegaga-2 …")
    pegaga = extract_pegaga(os.path.join(RES, "pegaga results.txt"))
    write_pegaga(pegaga, os.path.join(RES, "Pegaga2_ML_Dataset.xlsx"))

    print("\nDone.")


if __name__ == "__main__":
    main()
